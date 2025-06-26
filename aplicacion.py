import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, date
import re

st.set_page_config(page_title="Producción Yeseria", layout="wide")
st.image("logo.png", width=100)
st.title("📋 FORMULARIO PRODUCCIÓN DE YESERIA")

ruta_archivo = "BASE_FINAL.xlsx"


def limpiar_texto(texto):
    texto = texto.strip().upper()
    texto = texto.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    texto = re.sub(r"[^A-Z0-9]", "", texto)
    return texto

def obtener_nombre(codigo, df_operarios):
    fila = df_operarios[df_operarios["CÓDIGO"] == codigo]
    return fila.iloc[0]["OPERARIO"] if not fila.empty else "NO ENCONTRADO"

def obtener_hora_molde(molde, base_produccion):
    fila = base_produccion[base_produccion["COD MAT"].astype(str).str.strip().str.upper() == molde.strip().upper()]
    if not fila.empty:
        moldes_turno = fila.iloc[0]["MOLDES/TURNO"]
        personas_molde = fila.iloc[0]["PERSONAS/MOLDE"]
        if pd.notna(moldes_turno) and pd.notna(personas_molde) and moldes_turno != 0:
            return 8 * (personas_molde / moldes_turno)
    return None

@st.cache_data
def cargar_datos():
    try:
        base_produccion = pd.read_excel(ruta_archivo, sheet_name="Base_Produccion")
        tiempo_fallas = pd.read_excel(ruta_archivo, sheet_name="Tiempo_Fallas")
        operarios = pd.read_excel(ruta_archivo, sheet_name="Operarios", dtype={"CÓDIGO": str})
        return base_produccion, tiempo_fallas, operarios
    except Exception as e:
        st.error(f"Error al cargar los datos: {e}")
        return None, None, None

base_produccion, tiempo_fallas, operarios = cargar_datos()

if not all([base_produccion is not None, tiempo_fallas is not None, operarios is not None]):
    st.stop()

moldes = base_produccion["COD MAT"].dropna().astype(str).unique().tolist()
codigos_operarios = operarios["CÓDIGO"].dropna().astype(str).tolist()
piezas = tiempo_fallas["CODIGO"].dropna().astype(str).unique().tolist()
partes = tiempo_fallas["PARTE MOLDE"].dropna().astype(str).unique().tolist()

st.header("📝 Registro Producción Y Tiempos")

with st.form("formulario_final"):
    fecha = st.date_input("Fecha", max_value=date.today())
    molde = st.selectbox("Molde", options=[""] + moldes)
    cantidad_total = st.number_input("Cantidad Total Producida", min_value=0, value=0)

    st.subheader(" Ingrese Los Operarios que Trabajaron")
    operarios_merma = []

    for i in range(1, 6):
        st.markdown(f"### 👷Operario {i}")
        op_codigo = st.selectbox(f"Código Operario", options=[""] + codigos_operarios, key=f"op_{i}")

        st.markdown("##### 📄 Informe de pieza mal hecha")

        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
        with col1:
            pieza = st.selectbox(f"Pieza", options=[""] + piezas, key=f"pieza_{i}")
        with col2:
            parte = st.selectbox(f"Parte Molde", options=[""] + partes, key=f"parte_{i}")
        with col3:
            cant_merma = st.number_input(f"Cantidad", min_value=0, value=0, key=f"cant_{i}")
        with col4:
            cant_kg = st.number_input(f"Cantidad en KG", min_value=0.0, value=0.0, step=0.1, key=f"kg_{i}")

        operarios_merma.append({
            "codigo": op_codigo,
            "pieza": pieza,
            "parte": parte,
            "cantidad_merma": cant_merma,
            "cantidad_kg": cant_kg
        })

    submit = st.form_submit_button("✅ Guardar Registro de Producción")

if submit:
    import pytz
    zona_ecuador = pytz.timezone('America/Guayaquil')
    ahora = datetime.now(zona_ecuador)
    fecha_dt = datetime.combine(fecha, ahora.time())

    operadores = [op["codigo"] for op in operarios_merma if op["codigo"] != ""]
    hora = obtener_hora_molde(molde, base_produccion)

    fila_molde = base_produccion[base_produccion["COD MAT"].astype(str).str.strip().str.upper() == molde.strip().upper()]
    max_moldes = fila_molde.iloc[0]["MOLDES/TURNO"] if not fila_molde.empty else None

    if fecha is None:
        st.warning("⚠️ Debes ingresar la fecha.")
    elif molde == "":
        st.warning("⚠️ Debes seleccionar un molde.")
    elif fecha > datetime.now().date():
        st.warning("⚠️ La fecha no puede ser superior a hoy.")
    elif len(operadores) == 0:
        st.warning("⚠️ Debes ingresar al menos un código de operario.")
    elif len(operadores) != len(set(operadores)):
        st.warning("⚠️ No puede haber operarios con el mismo codigo")
    elif max_moldes is not None and cantidad_total > max_moldes:
        st.warning(f"⚠️ La cantidad supera el máximo permitido para el molde: {max_moldes}")
    elif hora is None:
        st.warning("⚠️ No se pudo calcular la hora por molde.")
    else:
        moldes_persona = cantidad_total / len(operadores)
        tiempo_usado = moldes_persona * hora
        indicador = round((tiempo_usado / 8) * 100, 1)

        if indicador > 100:
            st.warning("⚠️ Indicador supera el 100% por operario.")
        else:
            registros = []
            for op in operarios_merma:
                if op["codigo"] != "":
                    tiempo_merma = 0.0
                    if op["pieza"] in tiempo_fallas["CODIGO"].values:
                        tiempo_def = tiempo_fallas[tiempo_fallas["CODIGO"] == op["pieza"]]["TIEMPO (MIN)"].values
                        tiempo_merma = float(tiempo_def[0]) * op["cantidad_merma"] if len(tiempo_def) > 0 else 0.0

                    registros.append({
                        "Fecha": fecha_dt,
                        "Molde": molde,
                        "Moldes/Persona": round(moldes_persona, 2),
                        "Código": op["codigo"],
                        "Nombre": obtener_nombre(op["codigo"], operarios),
                        "Tiempo Usado": round(tiempo_usado, 2),
                        "Indicador de Producción": f"{indicador}%",
                        "Pieza": op["pieza"],
                        "Parte": op["parte"],
                        "Cantidad": op["cantidad_merma"],
                        "Cantidad en KG": round(op["cantidad_kg"], 2),
                        "Tiempo en Minutos": round(tiempo_merma, 2),
                        "Indicador de Tiempo": f"{round((tiempo_merma / 480) * 100, 2)}%"
                    })

            df_final = pd.DataFrame(registros)

            try:
                if os.path.exists(ruta_archivo):
                    libro = load_workbook(ruta_archivo)
                    if "FINAL" in libro.sheetnames:
                        existente = pd.read_excel(ruta_archivo, sheet_name="FINAL")
                        df_final = pd.concat([existente, df_final], ignore_index=True)

                with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_final.to_excel(writer, sheet_name="FINAL", index=False)
                st.success("✅ Registro Final guardado correctamente.")

            except Exception as e:
                st.error(f"❌ Error al guardar: {e}")

# Mostrar tabla FINAL y eliminar registros
try:
    libro = load_workbook(ruta_archivo)
    if "FINAL" in libro.sheetnames:
        df_final = pd.read_excel(ruta_archivo, sheet_name="FINAL")
        df_final = df_final.iloc[::-1]

        st.header("📊 REGISTROS DE PRODUCCIÓN")
        st.dataframe(df_final)

        st.subheader("🗑️ Eliminar Registro")
        if not df_final.empty:
            df_final["RESUMEN"] = df_final.apply(lambda row: f"{row['Fecha']} | {row['Molde']} | {row['Código']}", axis=1)
            indices = df_final.index.tolist()
            opciones = [f"{i} - {df_final.at[i, 'RESUMEN']}" for i in indices]

            seleccion = st.selectbox("Selecciona el registro a eliminar", options=opciones)
            index_a_eliminar = int(seleccion.split(" - ")[0])

            if st.button("Eliminar registro seleccionado"):
                df_final = df_final.drop(index=index_a_eliminar).reset_index(drop=True)
                with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_final.drop(columns=["RESUMEN"]).to_excel(writer, sheet_name="FINAL", index=False)
                st.success("✅ Registro eliminado correctamente.")
                st.rerun()
    else:
        st.info("ℹ️ No hay registros en la hoja 'FINAL'.")
except Exception as e:
    st.error(f"❌ Error mostrando registros: {e}")

# 🔍 Buscador de Producción Final
st.header("🔍 Buscador de Producción Real")

def cargar_datos_final(ruta):
    try:
        df = pd.read_excel(ruta, sheet_name="FINAL")
        return df
    except Exception as e:
        st.error(f"No se pudo cargar la hoja FINAL: {e}")
        return pd.DataFrame()

df_final = cargar_datos_final(ruta_archivo)

if df_final.empty:
    st.info("No hay datos en la hoja FINAL para mostrar.")
else:
    if not pd.api.types.is_datetime64_any_dtype(df_final['Fecha']):
        df_final['Fecha'] = pd.to_datetime(df_final['Fecha'], errors='coerce')

    codigos_disponibles = df_final["Código"].dropna().astype(str).unique().tolist()

    fecha_buscar = st.date_input("Selecciona la fecha", value=None, key="buscar_fecha")
    cod_operario_buscar = st.selectbox("Selecciona el código de operario", options=[""] + codigos_disponibles, key="buscar_codigo")

    df_filtrado = df_final.copy()
    if fecha_buscar:
        df_filtrado = df_filtrado[df_filtrado['Fecha'].dt.date == fecha_buscar]
    if cod_operario_buscar != "":
        df_filtrado = df_filtrado[df_filtrado["Código"].astype(str) == cod_operario_buscar]
    df_filtrado = df_filtrado.sort_values(by="Fecha", ascending=False)
    if df_filtrado.empty:
        st.warning("No se encontraron registros con los filtros aplicados.")
    else:
        def parse_percent(valor):
            if isinstance(valor, str) and valor.endswith('%'):
                return float(valor.strip('%'))
            try:
                return float(valor)
            except:
                return 0.0

        df_filtrado['Indicador Producción (%)'] = df_filtrado['Indicador de Producción'].apply(parse_percent)
        df_filtrado['Indicador Tiempo (%)'] = df_filtrado['Indicador de Tiempo'].apply(parse_percent)

        df_filtrado['Producción Real Trabajada'] = df_filtrado['Indicador Producción (%)'] - df_filtrado['Indicador Tiempo (%)']
        df_filtrado['Producción Real Trabajada'] = df_filtrado['Producción Real Trabajada'].apply(lambda x: f"{x:.2f}%")

        columnas_mostrar = [
            'Fecha', 'Molde', 'Moldes/Persona', 'Código', 'Nombre','Cantidad en KG',
            'Cantidad', 'Indicador de Producción', 'Indicador de Tiempo',
            'Producción Real Trabajada'
        ]

        columnas_mostrar = [col for col in columnas_mostrar if col in df_filtrado.columns]

        st.header("📊 Resultados de Producción Real Trabajada")
        st.dataframe(df_filtrado[columnas_mostrar].reset_index(drop=True))
