import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, date
import re
import streamlit.components.v1 as components

st.set_page_config(page_title="Producción Yeseria", layout="wide")

if st.session_state.get("__desplazar_temp", False):
    components.html(
        """
        <script>
            const streamlitDoc = window.parent.document;
            const rootNode = streamlitDoc.querySelector('section.main');
            if (rootNode) {
                rootNode.scrollTo({ top: 0, behavior: 'smooth' });
            } else {
                window.parent.scrollTo({ top: 0, behavior: 'smooth' });
            }
        </script>
        """,
        height=0,
    )
    del st.session_state["__desplazar_temp"]

st.image("logo.png", width=100)
st.title("📋 FORMULARIO PRODUCCIÓN DE YESERIA")

ruta_archivo = "BASE_FINAL.xlsx"

def limpiar_texto(texto):
    texto = texto.strip().upper()
    texto = texto.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    texto = re.sub(r"[^A-Z0-9]", "", texto)
    return texto

def obtener_nombre(codigo, df_operarios):
    fila = df_operarios[df_operarios["CÓDIGO"] == codigo]
    return fila.iloc[0]["OPERARIO"] if not fila.empty else "NO ENCONTRADO"

def obtener_hora_molde(molde, base_produccion):
    fila = base_produccion[base_produccion["COD MAT"].astype(str).str.strip().str.upper() == molde.strip().upper()]
    if not fila.empty:
        moldes_turno = fila.iloc[0]["MOLDES/TURNO"]
        personas_molde = fila.iloc[0]["PERSONAS/MOLDE"]
        if pd.notna(moldes_turno) and pd.notna(personas_molde) and moldes_turno != 0:
            return 8 * (personas_molde / moldes_turno)
    return None

@st.cache_data
def cargar_datos():
    try:
        base_produccion = pd.read_excel(ruta_archivo, sheet_name="Base_Produccion")
        tiempo_fallas = pd.read_excel(ruta_archivo, sheet_name="Tiempo_Fallas")
        operarios = pd.read_excel(ruta_archivo, sheet_name="Operarios", dtype={"CÓDIGO": str})

        tiempo_fallas.columns = tiempo_fallas.columns.str.strip()
        base_produccion.columns = base_produccion.columns.str.strip()
        operarios.columns = operarios.columns.str.strip()

        return base_produccion, tiempo_fallas, operarios
    except Exception as e:
        st.error(f"Error al cargar los datos: {e}")
        return None, None, None

def cargar_final():
    try:
        df = pd.read_excel(ruta_archivo, sheet_name="FINAL")
        if not pd.api.types.is_datetime64_any_dtype(df['Fecha']):
            df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
        return df
    except Exception as e:
        # Si la hoja no existe o hay error, devolver DF vacío
        return pd.DataFrame()

base_produccion, tiempo_fallas, operarios = cargar_datos()
df_final = cargar_final() 

if not all([base_produccion is not None, tiempo_fallas is not None, operarios is not None]):
    st.stop()

moldes = base_produccion["COD MAT"].dropna().astype(str).unique().tolist()
codigos_operarios = operarios["CÓDIGO"].dropna().astype(str).tolist()
partes = tiempo_fallas["PARTE MOLDE"].dropna().astype(str).str.strip().str.upper().unique().tolist()
lineas_disponibles = tiempo_fallas["LINEA"].dropna().astype(str).unique().tolist()

fecha = st.date_input("Fecha", value=st.session_state.get("fecha", date.today()), max_value=date.today(), key="fecha")

if "molde" not in st.session_state:
    st.session_state["molde"] = ""

lista_moldes = [""] + moldes
molde = st.selectbox("Molde", options=[""] + moldes, key="molde")
cantidad_total = st.number_input("Cantidad Total Producida", min_value=0, value=st.session_state.get("cantidad_total", 0), key="cantidad_total")

partes_por_letra = {
    "I": ["BASE", "TAPA", "LATERAL"],
    "D": ["MACHO", "HEMBRA"],
    "L": ["MACHO", "HEMBRA"],
    "V": ["MACHO", "HEMBRA"],
    "O": ["BASE", "TAPA", "LATERAL"],
    "T": ["MACHO", "HEMBRA"],
}

# Obtener las piezas relacionadas al molde seleccionado (filtradas desde tiempo_fallas)
piezas_disponibles = [molde] if molde else []

st.subheader("Ingreso Operarios")

with st.form("formulario_final"):
    operarios_merma = []

    for i in range(1, 6):
        with st.expander(f"👷 Operario {i}", expanded=(i == 1)):
            codigo_default = st.session_state.get(f"op_{i}", "")
            op_codigo = st.selectbox(
                f"Código Operario",
                options=[""] + codigos_operarios,
                index=([""] + codigos_operarios).index(codigo_default) if codigo_default in codigos_operarios else 0,
                key=f"op_{i}"
            )

            col1, col2, col3 = st.columns(3)

            with col1:
                st.markdown("#### 🛠️ Pieza Mal Hecha")
                pieza = molde if molde else ""
                st.text_input(f"Pieza", value=pieza, disabled=True, key=f"pieza_{i}")

            with col2:
                st.markdown("#### ")
                letra_inicial = molde[0].upper() if molde else ""
                partes_filtradas = partes_por_letra.get(letra_inicial, partes)
                parte_default = st.session_state.get(f"parte_{i}", "")
                parte = st.selectbox(
                    f"Parte Molde",
                    options=[""] + partes_filtradas,
                    index=([""] + partes_filtradas).index(parte_default) if parte_default in partes_filtradas else 0,
                    key=f"parte_{i}"
                )

            with col3:
                st.markdown("#### ")
                cantidad_input = st.number_input(
                    f"Cantidad",
                    min_value=0,
                    value=st.session_state.get(f"cant_{i}", 0),
                    step=1,
                    key=f"cant_{i}"
                )

            st.markdown("##### 🔄 Informe de Retrabajo")
            colr1, colr2, colr3 = st.columns([3, 3, 4])

            with colr1:
                molde_retra_default = st.session_state.get(f"molde_retrabajo_{i}", molde)
                molde_retrabajo = st.selectbox(
                    f"Molde Retrabajo",
                    options=[""] + moldes,
                    index=([""] + moldes).index(molde_retra_default) if molde_retra_default in moldes else 0,
                    key=f"molde_retrabajo_{i}"
                )

            with colr2:
                linea_default = st.session_state.get(f"linea_retrabajo_{i}", "")
                linea_retrabajo = st.selectbox(
                    "Línea",
                    options=[""] + lineas_disponibles,
                    index=([""] + lineas_disponibles).index(linea_default) if linea_default in lineas_disponibles else 0,
                    key=f"linea_retrabajo_{i}"
                )

            with colr3:
                col_horas, col_minutos = st.columns([1, 1])
                with col_horas:
                    horas_retrabajo = st.number_input(
                        "Horas",
                        min_value=0,
                        max_value=8,
                        value=st.session_state.get(f"horas_retrabajo_{i}", 0),
                        step=1,
                        key=f"horas_retrabajo_{i}"
                    )
                with col_minutos:
                    minutos_retrabajo = st.number_input(
                        "Minutos",
                        min_value=0,
                        max_value=59,
                        value=st.session_state.get(f"minutos_retrabajo_{i}", 0),
                        step=1,
                        key=f"minutos_retrabajo_{i}"
                    )

        tiempo_retrabajo_total = horas_retrabajo * 60 + minutos_retrabajo
        indicador_retrabajo = round((tiempo_retrabajo_total / 480) * 100, 2)

        operarios_merma.append({
            "codigo": op_codigo,
            "pieza": pieza,
            "parte": parte,
            "cantidad_merma": cantidad_input,
            "molde_retrabajo": molde_retrabajo,
            "linea_retrabajo": linea_retrabajo,
            "tiempo_retrabajo_min": tiempo_retrabajo_total,
            "indicador_retrabajo": indicador_retrabajo,
        })

    submit = st.form_submit_button("✅ Guardar Registro de Producción")


if st.button("🧹 Limpiar Formulario"):
    # Guardamos una bandera para activar el scroll
    st.session_state["__desplazar_temp"] = True

    # Guardamos claves que queremos conservar
    claves_conservar = ["__desplazar_temp"]

    # Borramos todas las demás claves (formulario completo)
    claves_a_borrar = [clave for clave in st.session_state.keys() if clave not in claves_conservar]
    for clave in claves_a_borrar:
        del st.session_state[clave]

    st.rerun()

if st.session_state.get("registro_exitoso", False):
    st.success("✅ Registro guardado con éxito.")
    del st.session_state["registro_exitoso"]

if submit:
    # Validar que ningún operario ya tenga un registro ese día
    for op in operarios_merma:
        if op["codigo"] != "":
            registros_existentes = df_final[
                (df_final["Código"].astype(str) == str(op["codigo"])) &
                (df_final["Fecha"].dt.date == fecha)
            ]
            if not registros_existentes.empty:
                st.warning(f"⚠️ El operario {op['codigo']} ya tiene un registro guardado para la fecha {fecha}. No se puede registrar más de una vez por día.")
                st.stop()

    for i, op in enumerate(operarios_merma, start=1):
        if op["tiempo_retrabajo_min"] > 480:
            st.warning(f"⚠️ Operario {i} ingresó más de 8 horas de retrabajo.")
            st.stop()

    for i, op in enumerate(operarios_merma, start=1):
        tiene_datos_merma = (op["pieza"] != "" and op["parte"] != "" and op["cantidad_merma"] > 0)
        tiene_datos_retrabajo = (op["molde_retrabajo"] != "" or op["linea_retrabajo"] != "" or op["tiempo_retrabajo_min"] > 0)

        if op["codigo"] and op["tiempo_retrabajo_min"] > 480 and (tiene_datos_merma or tiene_datos_retrabajo):
            st.warning(f"⚠️ El ingreso es más de 8 horas. No se permite superar las 8 horas.")
            st.stop()

    for i, op in enumerate(operarios_merma, start=1):
        # Validación para evitar datos incompletos en piezas mal hechas
        campos_llenos = any([
            op["parte"],
            op["cantidad_merma"] > 0,
            op["molde_retrabajo"],
            op["linea_retrabajo"],
            op["tiempo_retrabajo_min"] > 0
        ])
        if op["codigo"] and not campos_llenos:
            continue
        if not op["codigo"] and not campos_llenos:
            continue
        if campos_llenos and not op["codigo"]:
            st.warning(f"⚠️ El Operario {i} ingresó datos sin seleccionar un código.")
            st.stop()
        if (op["parte"] == "" and op["cantidad_merma"] > 0) or (op["parte"] != "" and op["cantidad_merma"] == 0):
            st.warning(f"⚠️ El Operario {i} tiene datos incompletos en piezas mal hechas.")
            st.stop()

        # NUEVA VALIDACIÓN PARA INFORME DE RETRABAJO
        molde_retrabajo = op["molde_retrabajo"]
        linea_retrabajo = op["linea_retrabajo"]
        tiempo_retrabajo_min = op["tiempo_retrabajo_min"]

        if molde_retrabajo != "":
            if linea_retrabajo == "" or tiempo_retrabajo_min == 0:
                st.warning(f"⚠️ El Operario {i} tiene datos incompletos en Molde Retrabajo.")
                st.stop()

        if (linea_retrabajo != "" or tiempo_retrabajo_min > 0) and molde_retrabajo == "":
            st.warning(f"⚠️ El Operario {i} tiene datos incompletos en Molde Retrabajo.")
            st.stop()

    if cantidad_total == 0:
        st.warning("⚠️ La cantidad total producida no puede ser 0. Por favor, ingresa un valor mayor a cero para continuar.")
    else:
        suma_merma = sum(op["cantidad_merma"] for op in operarios_merma)
        if suma_merma > cantidad_total:
            st.warning(f"⚠️ La suma total de piezas mal hechas ({suma_merma}) supera la cantidad producida ({cantidad_total}). Verifica los datos.")
            st.stop()

        ahora = datetime.now()
        fecha_dt = datetime.combine(fecha, ahora.time())

        operadores = [op["codigo"] for op in operarios_merma if op["codigo"] != ""]
        hora = obtener_hora_molde(molde, base_produccion)

        fila_molde = base_produccion[base_produccion["COD MAT"].astype(str).str.strip().str.upper() == molde.strip().upper()]
        max_moldes = fila_molde.iloc[0]["MOLDES/TURNO"] if not fila_molde.empty else None

        if fecha is None:
            st.warning("⚠️ Debes ingresar la fecha.")
        elif molde == "":
            st.warning("⚠️ Debes seleccionar un molde.")
        elif fecha > datetime.now().date():
            st.warning("⚠️ La fecha no puede ser superior a hoy.")
        elif len(operadores) == 0:
            st.warning("⚠️ Debes ingresar al menos un código de operario.")
        elif len(operadores) != len(set(operadores)):
            st.warning("⚠️ No puede haber operarios con el mismo código.")
        elif max_moldes is not None and cantidad_total > max_moldes:
            st.warning(f"⚠️ La cantidad supera el máximo permitido para el molde: {max_moldes}")
        elif hora is None:
            st.warning("⚠️ No se pudo calcular la hora por molde.")
        else:
            moldes_persona = cantidad_total / len(operadores)
            tiempo_usado = moldes_persona * hora
            indicador = round((tiempo_usado / 8) * 100, 1)

            if indicador > 100:
                st.warning("⚠️ Indicador supera el 100% por operario.")
            else:
                registros = []
                for op in operarios_merma:
                    if op["codigo"] != "":
                        tiempo_merma = 0.0
                        cantidad_kg = 0.0

                        pieza_clean = str(op["pieza"]).strip().upper()
                        parte_clean = str(op["parte"]).strip().upper()

                        df_tiempo_fallas = tiempo_fallas.copy()
                        df_tiempo_fallas["CODIGO"] = df_tiempo_fallas["CODIGO"].astype(str).str.strip().str.upper()
                        df_tiempo_fallas["PARTE MOLDE"] = df_tiempo_fallas["PARTE MOLDE"].astype(str).str.strip().str.upper()

                        fila_kg = df_tiempo_fallas[
                            (df_tiempo_fallas["CODIGO"] == pieza_clean) &
                            (df_tiempo_fallas["PARTE MOLDE"] == parte_clean)
                        ]

                        if not fila_kg.empty:
                            try:
                                kg_unitario = float(fila_kg.iloc[0]["CANTIDAD KG"])
                                cantidad_kg = kg_unitario * op["cantidad_merma"]
                            except:
                                cantidad_kg = 0.0

                            try:
                                tiempo_unitario = float(fila_kg.iloc[0]["TIEMPO (MIN)"])
                                tiempo_merma = tiempo_unitario * op["cantidad_merma"]
                            except:
                                tiempo_merma = 0.0

                        pieza_final = op["pieza"]
                        if op["parte"] == "" or op["cantidad_merma"] == 0:
                            pieza_final = None

                        registros.append({
                            "Fecha": fecha_dt,
                            "Molde": molde,
                            "Moldes/Persona": round(moldes_persona, 2),
                            "Código": op["codigo"],
                            "Nombre": obtener_nombre(op["codigo"], operarios),
                            "Tiempo Usado": round(tiempo_usado, 2),
                            "Indicador de Producción": f"{indicador}%",
                            "Pieza": pieza_final,
                            "Parte": op["parte"],
                            "Cantidad": op["cantidad_merma"],
                            "Cantidad KG": cantidad_kg,
                            "Tiempo en Minutos": round(tiempo_merma, 2),
                            "Indicador de Tiempo": f"{round((tiempo_merma / 480) * 100, 2)}%",
                            "Molde Retrabajo": op["molde_retrabajo"],
                            "Linea Retrabajo": op["linea_retrabajo"],
                            "Tiempo Retrabajo (minutos)": op["tiempo_retrabajo_min"],
                            "Indicador Retrabajo": f"{op['indicador_retrabajo']}%"
                        })

                df_final = pd.DataFrame(registros)

                try:
                    if os.path.exists(ruta_archivo):
                        libro = load_workbook(ruta_archivo)
                        if "FINAL" in libro.sheetnames:
                            existente = pd.read_excel(ruta_archivo, sheet_name="FINAL")
                            df_final = pd.concat([existente, df_final], ignore_index=True)

                    with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_final.to_excel(writer, sheet_name="FINAL", index=False)
                    st.session_state["registro_exitoso"] = True
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Error al guardar: {e}")


# Función para cargar y ordenar los registros por fecha descendente
def cargar_final_ordenado(ruta):
    try:
        df = pd.read_excel(ruta, sheet_name="FINAL")
        if not pd.api.types.is_datetime64_any_dtype(df['Fecha']):
            df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
        return df.sort_values(by="Fecha", ascending=False).reset_index(drop=True)
    except Exception as e:
        st.error(f"No se pudo cargar la hoja FINAL: {e}")
        return pd.DataFrame()

# Mostrar tabla FINAL y eliminar registros
try:
    df_final = cargar_final_ordenado(ruta_archivo)

    if not df_final.empty:
        st.header("📊 REGISTROS DE PRODUCCIÓN")
        st.dataframe(df_final)

        st.subheader("🗑️ Eliminar Registro")
        df_final["RESUMEN"] = df_final.apply(lambda row: f"{row['Fecha']} | {row['Código']}", axis=1)
        indices = df_final.index.tolist()
        opciones = [f"{i} - {df_final.at[i, 'RESUMEN']}" for i in indices]

        seleccion = st.selectbox("Selecciona el registro a eliminar", options=opciones)
        index_a_eliminar = int(seleccion.split(" - ")[0])

        if st.button("Eliminar registro seleccionado"):
            # Obtener la fecha y hora exacta del registro seleccionado
            fecha_objetivo = df_final.at[index_a_eliminar, "Fecha"]

            # Filtrar todos los registros con la misma fecha y hora exacta
            df_final = df_final[df_final["Fecha"] != fecha_objetivo].reset_index(drop=True)

            # Guardar nuevamente el archivo sin los registros eliminados
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_final.drop(columns=["RESUMEN"]).to_excel(writer, sheet_name="FINAL", index=False)

            st.success(f"✅ Todos los registros con fecha {fecha_objetivo} fueron eliminados correctamente.")
            st.session_state.pop("activar_filtro", None)
            st.rerun()
    else:
        st.info("ℹ️ No hay registros en la hoja 'FINAL'.")
except Exception as e:
    st.error(f"❌ Error mostrando registros: {e}")

# 🔍 Buscador de Producción Final
st.header("🔍 Buscador de Producción Real")

def cargar_datos_final(ruta):
    try:
        df = pd.read_excel(ruta, sheet_name="FINAL")
        return df
    except Exception as e:
        st.error(f"No se pudo cargar la hoja FINAL: {e}")
        return pd.DataFrame()

df_final = cargar_datos_final(ruta_archivo)

if df_final.empty:
    st.info("No hay datos en la hoja FINAL para mostrar.")
else:
    if not pd.api.types.is_datetime64_any_dtype(df_final['Fecha']):
        df_final['Fecha'] = pd.to_datetime(df_final['Fecha'], errors='coerce')

    codigos_disponibles = df_final["Código"].dropna().astype(str).unique().tolist()
    activar_filtro = st.checkbox("🔍 Aplicar filtro por fecha y código", key="activar_filtro")

    fecha_inicio = None
    fecha_fin = None
    cod_operario_buscar = ""
    filtros_validos = True
    mostrar_tabla = False

    df_filtrado = df_final.copy()

    if activar_filtro:
        colf1, colf2 = st.columns(2)
        with colf1:
            fecha_inicio = st.date_input("📆 Fecha inicial", key="buscar_fecha_inicio")
        with colf2:
            fecha_fin = st.date_input("📆 Fecha final", key="buscar_fecha_fin")

        if fecha_inicio > fecha_fin:
            st.warning("⚠️ La fecha inicial no puede ser mayor que la fecha final.")
            filtros_validos = False
        else:
            cod_operario_buscar = st.selectbox("👷 Código de operario", options=[""] + codigos_disponibles, key="buscar_codigo")

            if fecha_inicio and fecha_fin and cod_operario_buscar != "":
                mostrar_tabla = True
    else:
        mostrar_tabla = True

    if mostrar_tabla:
        if activar_filtro and filtros_validos:
            df_filtrado = df_filtrado[
                (df_filtrado['Fecha'].dt.date >= fecha_inicio) &
                (df_filtrado['Fecha'].dt.date <= fecha_fin) &
                (df_filtrado["Código"].astype(str) == cod_operario_buscar)
            ]

        df_filtrado = df_filtrado.sort_values(by="Fecha", ascending=False)

        if df_filtrado.empty:
            st.warning("No se encontraron registros con los filtros aplicados.")
        else:
            def parse_percent(valor):
                if isinstance(valor, str) and valor.endswith('%'):
                    return float(valor.strip('%'))
                try:
                    return float(valor)
                except:
                    return 0.0

            df_filtrado['Indicador Producción (%)'] = df_filtrado['Indicador de Producción'].apply(parse_percent)
            df_filtrado['Indicador Tiempo (%)'] = df_filtrado['Indicador de Tiempo'].apply(parse_percent)
            df_filtrado['Indicador Retrabajo (%)'] = df_filtrado['Indicador Retrabajo'].apply(parse_percent)

            df_filtrado['Producción Real Trabajada Num'] = (
                df_filtrado['Indicador Producción (%)']
                - df_filtrado['Indicador Tiempo (%)']
                - df_filtrado['Indicador Retrabajo (%)']
            )

            df_filtrado['Producción Real Trabajada'] = df_filtrado['Producción Real Trabajada Num'].apply(lambda x: f"{x:.2f}%")

            columnas_mostrar = [
                'Fecha', 'Molde', 'Moldes/Persona', 'Código', 'Nombre',
                'Cantidad', 'Indicador de Producción', 'Indicador de Tiempo',
                'Indicador Retrabajo','Producción Real Trabajada'
            ]
            columnas_mostrar = [col for col in columnas_mostrar if col in df_filtrado.columns]

            st.header("📊 Resultados de Producción Real Trabajada")
            st.dataframe(df_filtrado[columnas_mostrar].reset_index(drop=True))

            # Aquí calculamos y mostramos el promedio simple y el porcentaje ponderado real trabajado
            # Aquí calculamos y mostramos el promedio simple y el porcentaje ponderado real trabajado
            if activar_filtro and not df_filtrado.empty:
                dias_unicos = df_filtrado['Fecha'].dt.date.nunique()
                promedio_simple = df_filtrado['Producción Real Trabajada Num'].mean()
                horas_trabajadas_reales = (df_filtrado['Producción Real Trabajada Num'] / 100 * 8).sum()
                horas_posibles = dias_unicos * 8
                porcentaje_ponderado = (horas_trabajadas_reales / horas_posibles) * 100

                st.markdown(f"### ✅ Promedio Producción Real Trabajada: **{promedio_simple:.2f}%**")
